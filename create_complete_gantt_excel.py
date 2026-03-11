"""
Generate Complete Study Gantt Excel
====================================
Creates a comprehensive Excel Gantt chart with ALL subsections from the study agenda.

Features:
- All sections and subsections from the agenda
- Effort (hours) column
- Progress (%) column  
- Notes column
- Color-coded weeks for activities
- Conditional formatting

Quick Test:
    python create_complete_gantt_excel.py
    # Opens study_plan_complete.xlsx

Dependencies:
    pip install openpyxl pandas
"""

from pathlib import Path
import webbrowser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule


def get_full_schedule():
    """
    Complete schedule data matching the 28-week study agenda.
    Returns list of: (Section, Subsection, Priority, Hours, Weeks_Active, Activity_Type)
    """
    
    # Define all items with their schedules
    # Format: (Section, Subsection, Priority, EstHours, [(week, activity_type), ...])
    
    data = [
        # ===== PHASE 0: Time Series (Weeks 1-2) =====
        ("4. Time Series", "4.1 Fundamentals", "[C]", 6, [(2, "R")]),
        ("4. Time Series", "4.2 Classical Methods", "[H]", 6, [(2, "R")]),
        ("4. Time Series", "4.3 Modern Methods", "[H]", 8, [(1, "S")]),
        ("4. Time Series", "4.4 Practical Considerations", "[H]", 8, [(1, "S")]),
        ("4. Time Series", "4.5 Uncertainty Quantification", "[M]", 6, [(2, "S")]),
        
        # ===== PHASE 1: Stats + ML Foundations (Weeks 3-7) =====
        ("2. Stats", "2.1 Probability Basics", "[H]", 8, [(3, "S")]),
        ("2. Stats", "2.2 Distributions", "[H]", 6, [(4, "S")]),
        ("2. Stats", "2.3 Statistical Inference", "[H]", 8, [(4, "S")]),
        ("2. Stats", "2.4 Advanced Theory", "[M]", 6, [(5, "S")]),
        
        ("3. ML/DL", "3.1.1 Core Concepts", "[C]", 4, [(4, "S")]),
        ("3. ML/DL", "3.1.2 Linear Models", "[C]", 6, [(4, "S")]),
        ("3. ML/DL", "3.1.3 Tree-Based Methods", "[C]", 8, [(5, "S")]),
        ("3. ML/DL", "3.1.4 Model Evaluation", "[C]", 6, [(5, "S")]),
        ("3. ML/DL", "3.1.5 SVM", "[M]", 4, [(6, "S")]),
        ("3. ML/DL", "3.1.6 Imbalanced Learning", "[H]", 6, [(6, "S")]),
        
        ("9. Breadth", "9.3 Survival Analysis", "[M]", 6, [(6, "S")]),
        
        ("Review", "Stats + ML Review", "-", 8, [(7, "R")]),
        ("Mock", "#1 Stats + ML Fundamentals", "-", 2, [(7, "M")]),
        
        # ===== PHASE 2: Causal Deep Dive (Weeks 8-13) =====
        ("1. Causal", "1.1.1 Potential Outcomes (Rubin)", "[C]", 6, [(8, "S")]),
        ("1. Causal", "1.1.2 Structural Causal Models (Pearl)", "[C]", 6, [(8, "S")]),
        ("1. Causal", "1.1.3 Identification Strategies", "[C]", 4, [(8, "S")]),
        
        ("1. Causal", "1.2.1 Randomized Experiments", "[C]", 8, [(9, "S")]),
        ("1. Causal", "1.2.2 Difference-in-Differences", "[C]", 6, [(9, "S")]),
        ("1. Causal", "1.2.3 Regression Discontinuity", "[H]", 4, [(9, "S")]),
        
        ("1. Causal", "1.2.4 Instrumental Variables", "[C]", 6, [(10, "S")]),
        ("1. Causal", "1.2.5 Synthetic Control", "[C]", 6, [(10, "S")]),
        
        ("1. Causal", "1.2.6 Propensity Score Methods", "[C]", 8, [(11, "S")]),
        ("1. Causal", "1.2.7 Bandits", "[H]", 6, [(11, "S")]),
        
        ("1. Causal", "1.3.1 Double ML (DML)", "[H]", 6, [(12, "S")]),
        ("1. Causal", "1.3.2 Causal Forests", "[H]", 4, [(12, "S")]),
        ("1. Causal", "1.3.3 Meta-Learners", "[H]", 6, [(12, "S")]),
        
        ("1. Causal", "1.4 Libraries (DoWhy, EconML)", "[H]", 6, [(13, "S")]),
        ("1. Causal", "1.5 Sensitivity Analysis", "[H]", 4, [(13, "S")]),
        
        ("Mock", "#2 Causal Deep Dive", "-", 2, [(13, "M")]),
        
        # ===== PHASE 3: DL + Project 1 (Weeks 14-18) =====
        ("3. ML/DL", "3.3.1 NN Fundamentals", "[H]", 6, [(14, "S")]),
        ("3. ML/DL", "3.3.2 Practical Considerations", "[M]", 4, [(14, "S")]),
        ("3. ML/DL", "3.3.3 Architectures Overview", "[H]", 4, [(15, "S")]),
        ("3. ML/DL", "3.3.4 Transfer Learning", "[M]", 4, [(15, "S")]),
        
        ("3. ML/DL", "3.2 Unsupervised Learning", "[M]", 6, [(16, "S")]),
        ("3. ML/DL", "3.4 Feature Engineering", "[H]", 6, [(16, "S")]),
        ("3. ML/DL", "3.5 Model Selection & Tuning", "[H]", 6, [(17, "S")]),
        
        ("1. Causal", "1.6 Applications", "[H]", 6, [(18, "S")]),
        
        ("Project", "Project 1: Causal Uplift", "-", 60, [(15, "P"), (16, "P"), (17, "P"), (18, "P")]),
        ("Mock", "#3 Full Causal + ML", "-", 2, [(18, "M")]),
        
        # ===== PHASE 4: NLP/GenAI + Project 2 (Weeks 19-23) =====
        ("5. NLP/GenAI", "5.1 NLP Fundamentals", "[M]", 4, [(19, "S")]),
        ("5. NLP/GenAI", "5.2 Transformers & LLMs", "[H]", 8, [(19, "S")]),
        ("5. NLP/GenAI", "5.3 RAG", "[H]", 8, [(20, "S")]),
        ("5. NLP/GenAI", "5.4 Agentic AI", "[M]", 6, [(21, "S")]),
        ("5. NLP/GenAI", "5.5 Causal + AI", "[M]", 4, [(21, "S")]),
        
        ("9. Breadth", "9.1 Markov Models / HMM", "[M]", 4, [(22, "S")]),
        ("9. Breadth", "9.2 RL Basics", "[L]", 3, [(22, "S")]),
        
        ("7. Coding", "7.3 Best Practices", "[M]", 4, [(23, "S")]),
        
        ("Project", "Project 2: TBD", "-", 50, [(21, "P"), (22, "P"), (23, "P")]),
        ("Mock", "#4 GenAI + System", "-", 2, [(23, "M")]),
        
        # ===== PHASE 5: Interview Sprint (Weeks 24-28) =====
        ("6. System Design", "6.1 Design Patterns", "[H]", 6, [(24, "S")]),
        ("6. System Design", "6.2 Key Components", "[M]", 6, [(24, "S")]),
        ("6. System Design", "6.3 Case Studies", "[H]", 8, [(25, "S")]),
        
        ("7. Coding", "7.1 Core Patterns", "[H]", 8, [(26, "S")]),
        ("7. Coding", "7.2 Python for ML", "[H]", 6, [(26, "S")]),
        
        ("8. Product Sense", "8.1 Problem Formulation", "[C]", 6, [(27, "S")]),
        ("8. Product Sense", "8.2 Stakeholder Communication", "[H]", 4, [(27, "S")]),
        ("8. Product Sense", "8.3 Case Study Practice", "[H]", 6, [(27, "S")]),
        
        ("Mock", "#5-8 Weekly Mocks", "-", 8, [(24, "M"), (25, "M"), (26, "M"), (27, "M"), (28, "M")]),
        ("Review", "Final Review", "-", 10, [(28, "R")]),
        
        # ===== PARALLEL: DSA =====
        ("DSA", "Daily Practice (30min)", "-", 84, [(w, "D") for w in range(1, 29)]),
    ]
    
    return data


def create_excel(data, output_path):
    """Create the Excel workbook with Gantt chart."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Study Gantt"
    
    # Define colors
    colors = {
        "S": "92D050",  # Green - Study
        "R": "00B0F0",  # Blue - Review
        "P": "FFC000",  # Orange - Project
        "M": "7030A0",  # Purple - Mock
        "D": "A6A6A6",  # Gray - DSA
    }
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Section", "Subsection", "Priority", "Hours", "Progress", "Notes"]
    week_headers = [f"W{w}" for w in range(1, 29)]
    all_headers = headers + week_headers
    
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Data rows
    for row_idx, (section, subsection, priority, hours, week_activities) in enumerate(data, 2):
        # Basic columns
        ws.cell(row=row_idx, column=1, value=section).border = thin_border
        ws.cell(row=row_idx, column=2, value=subsection).border = thin_border
        ws.cell(row=row_idx, column=3, value=priority).border = thin_border
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4, value=hours).border = thin_border
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=5, value="0%").border = thin_border
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6, value="").border = thin_border
        
        # Week columns
        for week in range(1, 29):
            col = 6 + week
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            
            # Check if this week has an activity
            for w, activity in week_activities:
                if w == week:
                    cell.value = activity
                    cell.fill = PatternFill(
                        start_color=colors.get(activity, "FFFFFF"),
                        end_color=colors.get(activity, "FFFFFF"),
                        fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal="center")
                    if activity in ["S", "R", "M"]:
                        cell.font = Font(bold=True, color="FFFFFF")
    
    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 25
    for col in range(7, 35):
        ws.column_dimensions[get_column_letter(col)].width = 4
    
    # Freeze panes
    ws.freeze_panes = "G2"
    
    # Add legend sheet
    legend_ws = wb.create_sheet("Legend")
    legend_data = [
        ("Code", "Activity", "Color"),
        ("S", "Study", "Green"),
        ("R", "Review", "Blue"),
        ("P", "Project", "Orange"),
        ("M", "Mock Interview", "Purple"),
        ("D", "DSA Practice", "Gray"),
    ]
    for row_idx, (code, activity, color) in enumerate(legend_data, 1):
        legend_ws.cell(row=row_idx, column=1, value=code)
        legend_ws.cell(row=row_idx, column=2, value=activity)
        legend_ws.cell(row=row_idx, column=3, value=color)
        if row_idx > 1:
            legend_ws.cell(row=row_idx, column=1).fill = PatternFill(
                start_color=colors.get(code, "FFFFFF"),
                end_color=colors.get(code, "FFFFFF"),
                fill_type="solid"
            )
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_data = [
        ("Metric", "Value"),
        ("Total Weeks", 28),
        ("Total Hours", "~560h"),
        ("Weekly Hours", "20h"),
        ("Primary Spike", "Causal Inference (~100h)"),
        ("Secondary Spike", "NLP/GenAI (~40h)"),
        ("Projects", 2),
        ("Mock Interviews", "7+"),
    ]
    for row_idx, (metric, value) in enumerate(summary_data, 1):
        summary_ws.cell(row=row_idx, column=1, value=metric)
        summary_ws.cell(row=row_idx, column=2, value=value)
    
    wb.save(output_path)
    return output_path


def main():
    """Generate the complete Excel Gantt chart."""
    print("Generating Complete Study Gantt Excel...")
    
    data = get_full_schedule()
    output_path = Path(__file__).parent / "study_plan_complete.xlsx"
    
    create_excel(data, output_path)
    
    print(f"Excel saved to: {output_path}")
    print(f"Total items: {len(data)}")
    print("Opening file...")
    
    # Open with default application
    import os
    os.startfile(output_path)


if __name__ == "__main__":
    main()
