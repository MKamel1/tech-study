"""
Generate COMPLETE Study Gantt Excel - All 121 Subsections
==========================================================
Parses the study agenda and creates an Excel with EVERY subsection.

Quick Test:
    python create_full_gantt_excel.py
"""

from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_agenda(agenda_path):
    """
    Parse the study agenda to extract all sections and subsections.
    Returns list of (section_num, section_name, subsection_id, subsection_name, priority)
    """
    with open(agenda_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    items = []
    current_section_num = ""
    current_section_name = ""
    current_subsection_id = ""
    current_subsection_name = ""
    
    lines = content.split('\n')
    
    for line in lines:
        # Match section headers like "## 1. Causal Inference"
        section_match = re.match(r'^## (\d+)\. (.+)$', line)
        if section_match:
            current_section_num = section_match.group(1)
            current_section_name = section_match.group(2).strip()
            continue
        
        # Match subsection headers like "### 1.1 Foundations **[C]**"
        subsection_match = re.match(r'^### (\d+\.\d+) (.+?) \*\*\[([CHML]|OPTIONAL)\]\*\*', line)
        if subsection_match:
            current_subsection_id = subsection_match.group(1)
            current_subsection_name = subsection_match.group(2).strip()
            continue
        
        # Match items like "- [ ] 1.1.1 Potential Outcomes Framework (Rubin) **[C]**"
        item_match = re.match(r'^- \[ \] (\d+\.\d+\.\d+) (.+?) \*\*\[([CHML]|OPTIONAL)\]\*\*', line)
        if item_match:
            item_id = item_match.group(1)
            item_name = item_match.group(2).strip()
            priority = item_match.group(3)
            
            items.append({
                'section_num': current_section_num,
                'section_name': current_section_name,
                'subsection_id': current_subsection_id,
                'subsection_name': current_subsection_name,
                'item_id': item_id,
                'item_name': item_name,
                'priority': f'[{priority}]'
            })
    
    return items


def get_week_schedule():
    """
    Define which weeks each section is studied.
    Returns dict: section_num -> [(start_week, end_week, activity_type), ...]
    """
    return {
        "1": [(8, 13, "S"), (18, 18, "S")],  # Causal: Weeks 8-13, 18
        "2": [(3, 5, "S"), (7, 7, "R")],      # Stats: Weeks 3-5, review 7
        "3": [(4, 6, "S"), (14, 17, "S")],    # ML/DL: Weeks 4-6, 14-17
        "4": [(1, 2, "S")],                    # Time Series: Weeks 1-2
        "5": [(19, 21, "S")],                  # NLP: Weeks 19-21
        "6": [(24, 25, "S")],                  # System Design: Weeks 24-25
        "7": [(23, 23, "S"), (26, 26, "S")],  # Coding: Weeks 23, 26
        "8": [(27, 27, "S")],                  # Product Sense: Week 27
        "9": [(6, 6, "S"), (22, 22, "S")],    # Breadth: Weeks 6, 22
        "10": [],                              # Optional - no scheduled weeks
    }


def estimate_hours(priority):
    """Estimate hours based on priority."""
    hour_map = {
        "[C]": 6,
        "[H]": 4,
        "[M]": 3,
        "[L]": 2,
        "[OPTIONAL]": 2,
    }
    return hour_map.get(priority, 3)


def create_excel(items, week_schedule, output_path):
    """Create the Excel workbook."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Study Gantt"
    
    # Colors
    colors = {
        "S": "92D050",  # Green - Study
        "R": "00B0F0",  # Blue - Review
        "P": "FFC000",  # Orange - Project
        "M": "7030A0",  # Purple - Mock
    }
    
    priority_colors = {
        "[C]": "FF6B6B",  # Red
        "[H]": "FFA500",  # Orange
        "[M]": "FFD93D",  # Yellow
        "[L]": "6BCB77",  # Green
        "[OPTIONAL]": "CCCCCC",  # Gray
    }
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Section", "Subsection", "Item ID", "Item Name", "Priority", "Hours", "Progress", "Notes"]
    week_headers = [f"W{w}" for w in range(1, 29)]
    all_headers = headers + week_headers
    
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    
    # Data rows
    current_section = ""
    for row_idx, item in enumerate(items, 2):
        section_display = ""
        if item['section_num'] != current_section:
            section_display = f"{item['section_num']}. {item['section_name']}"
            current_section = item['section_num']
        
        # Basic columns
        ws.cell(row=row_idx, column=1, value=section_display).border = thin_border
        ws.cell(row=row_idx, column=2, value=f"{item['subsection_id']} {item['subsection_name']}").border = thin_border
        ws.cell(row=row_idx, column=3, value=item['item_id']).border = thin_border
        ws.cell(row=row_idx, column=4, value=item['item_name']).border = thin_border
        
        # Priority with color
        priority_cell = ws.cell(row=row_idx, column=5, value=item['priority'])
        priority_cell.border = thin_border
        priority_cell.alignment = Alignment(horizontal="center")
        if item['priority'] in priority_colors:
            priority_cell.fill = PatternFill(
                start_color=priority_colors[item['priority']],
                end_color=priority_colors[item['priority']],
                fill_type="solid"
            )
        
        # Hours estimate
        hours = estimate_hours(item['priority'])
        ws.cell(row=row_idx, column=6, value=hours).border = thin_border
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")
        
        # Progress
        ws.cell(row=row_idx, column=7, value="0%").border = thin_border
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center")
        
        # Notes
        ws.cell(row=row_idx, column=8, value="").border = thin_border
        
        # Week columns - fill based on section schedule
        section_num = item['section_num']
        if section_num in week_schedule:
            for start, end, activity in week_schedule[section_num]:
                for week in range(start, end + 1):
                    col = 8 + week
                    cell = ws.cell(row=row_idx, column=col)
                    cell.value = activity
                    cell.fill = PatternFill(
                        start_color=colors.get(activity, "FFFFFF"),
                        end_color=colors.get(activity, "FFFFFF"),
                        fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # Add borders to all week cells
        for week in range(1, 29):
            ws.cell(row=row_idx, column=8 + week).border = thin_border
    
    # Add special rows for Projects, Mocks, DSA
    special_rows = [
        ("Project", "Project 1", "P1", "Causal Uplift Modeling", "-", 60, [(15, 18, "P")]),
        ("Project", "Project 2", "P2", "TBD (Choose at Week 16)", "-", 50, [(21, 23, "P")]),
        ("Mock", "Mock Interview", "M1", "#1 Stats + ML Fundamentals", "-", 2, [(7, 7, "M")]),
        ("Mock", "Mock Interview", "M2", "#2 Causal Deep Dive", "-", 2, [(13, 13, "M")]),
        ("Mock", "Mock Interview", "M3", "#3 Full Causal + ML", "-", 2, [(18, 18, "M")]),
        ("Mock", "Mock Interview", "M4", "#4 GenAI + System", "-", 2, [(23, 23, "M")]),
        ("Mock", "Mock Interview", "M5-8", "#5-8 Weekly Mocks", "-", 8, [(24, 28, "M")]),
        ("Review", "Final Review", "R1", "Full Review Week", "-", 10, [(28, 28, "R")]),
        ("DSA", "Daily Practice", "DSA", "30min daily throughout", "-", 84, [(1, 28, "S")]),
    ]
    
    row_idx = len(items) + 2
    for section, subsection, item_id, item_name, priority, hours, weeks in special_rows:
        ws.cell(row=row_idx, column=1, value=section).border = thin_border
        ws.cell(row=row_idx, column=2, value=subsection).border = thin_border
        ws.cell(row=row_idx, column=3, value=item_id).border = thin_border
        ws.cell(row=row_idx, column=4, value=item_name).border = thin_border
        ws.cell(row=row_idx, column=5, value=priority).border = thin_border
        ws.cell(row=row_idx, column=6, value=hours).border = thin_border
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=7, value="0%").border = thin_border
        ws.cell(row=row_idx, column=8, value="").border = thin_border
        
        for start, end, activity in weeks:
            for week in range(start, end + 1):
                col = 8 + week
                cell = ws.cell(row=row_idx, column=col)
                cell.value = activity
                cell.fill = PatternFill(
                    start_color=colors.get(activity, "FFFFFF"),
                    end_color=colors.get(activity, "FFFFFF"),
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True, color="FFFFFF")
        
        for week in range(1, 29):
            ws.cell(row=row_idx, column=8 + week).border = thin_border
        
        row_idx += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 20
    for col in range(9, 37):
        ws.column_dimensions[get_column_letter(col)].width = 4
    
    # Freeze panes
    ws.freeze_panes = "I2"
    
    # Add Legend sheet
    legend_ws = wb.create_sheet("Legend")
    legend_data = [
        ("Activity Codes", "", ""),
        ("S", "Study", "Green"),
        ("R", "Review", "Blue"),
        ("P", "Project", "Orange"),
        ("M", "Mock Interview", "Purple"),
        ("", "", ""),
        ("Priority Colors", "", ""),
        ("[C]", "Critical", "Red"),
        ("[H]", "High", "Orange"),
        ("[M]", "Medium", "Yellow"),
        ("[L]", "Low", "Green"),
        ("[OPTIONAL]", "Optional", "Gray"),
    ]
    for row_idx, row_data in enumerate(legend_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            legend_ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add Summary sheet
    summary_ws = wb.create_sheet("Summary")
    total_hours = sum(estimate_hours(item['priority']) for item in items)
    summary_data = [
        ("Study Agenda Summary", ""),
        ("", ""),
        ("Total Subsections", len(items)),
        ("Total Estimated Hours (topics)", total_hours),
        ("Project Hours", 110),
        ("DSA Hours", 84),
        ("Mock Interview Hours", 16),
        ("Review Hours", 10),
        ("Grand Total", total_hours + 110 + 84 + 16 + 10),
        ("", ""),
        ("Timeline", "28 weeks"),
        ("Weekly Hours", "20h"),
        ("", ""),
        ("Sections by Count", ""),
    ]
    
    # Count by section
    section_counts = {}
    for item in items:
        sec = item['section_num']
        section_counts[sec] = section_counts.get(sec, 0) + 1
    
    for sec in sorted(section_counts.keys(), key=int):
        item = next(i for i in items if i['section_num'] == sec)
        summary_data.append((f"Section {sec}: {item['section_name']}", section_counts[sec]))
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            summary_ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(output_path)
    return len(items)


def main():
    """Generate the complete Excel Gantt chart."""
    print("Generating COMPLETE Study Gantt Excel...")
    
    agenda_path = Path(__file__).parent / "as_study_agenda.md"
    output_path = Path(__file__).parent / "study_plan_complete.xlsx"
    
    items = parse_agenda(agenda_path)
    print(f"Parsed {len(items)} subsections from agenda")
    
    # Print by section
    section_counts = {}
    for item in items:
        sec = item['section_num']
        section_counts[sec] = section_counts.get(sec, 0) + 1
    
    print("\nSubsections per section:")
    for sec in sorted(section_counts.keys(), key=int):
        print(f"  Section {sec}: {section_counts[sec]} items")
    
    week_schedule = get_week_schedule()
    count = create_excel(items, week_schedule, output_path)
    
    print(f"\nExcel saved to: {output_path}")
    print(f"Total rows: {count} subsections + 9 special rows (projects, mocks, DSA)")
    print("\nOpening file...")
    
    import os
    os.startfile(output_path)


if __name__ == "__main__":
    main()
